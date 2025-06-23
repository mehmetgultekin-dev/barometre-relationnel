import streamlit as st
import pandas as pd
import openpyxl
import tempfile
import os
import json
from datetime import datetime
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode, ColumnsAutoSizeMode
from dotenv import load_dotenv
import hashlib
import io
import zipfile

import streamlit as st
import hashlib

# R√©cup√©ration s√©curis√©e depuis les secrets Streamlit
USERNAME = st.secrets["auth"]["username"]
PASSWORD_HASH = st.secrets["auth"]["password_hash"]

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.title("üîí Connexion s√©curis√©e")
    login = st.text_input("Identifiant")
    password = st.text_input("Mot de passe", type="password")
    if st.button("Se connecter"):
        if login == USERNAME and hash_password(password) == PASSWORD_HASH:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Identifiants incorrects. Veuillez r√©essayer.")
    st.stop()



# === LOGIQUE M√âTIER ==========================================================
class AnalyseRelationnelle:
    """Classe utilitaire pour calculer la vigilance d‚Äôune relation."""
    def __init__(self, relations_saisies):
        self.relations = relations_saisies

    @staticmethod
    def classer_relation(p_plus: int, p_moins: int) -> str:
        """
        Classifie une relation en fonction des scores positifs (p_plus) et n√©gatifs (p_moins).
        Met √† jour la logique pour "Positif pur" (3 P+) et introduit "Positif" (1 ou 2 P+),
        ainsi que "N√©gatif pur" (3 P-) et "N√©gatif" (1 ou 2 P-).
        """
        if p_plus == 3 and p_moins == 0:
            return "Positif pur" # R√©serv√© aux trois pics positifs (P+, I+, C+)
        elif p_plus == 0 and p_moins == 3:
            return "N√©gatif pur" # R√©serv√© aux trois pics n√©gatifs (P-, I-, C-)
        elif (p_plus == 1 or p_plus == 2) and p_moins == 0:
            return "Positif" # Pour un ou deux pics positifs, sans n√©gatif
        elif p_plus == 0 and (p_moins == 1 or p_moins == 2):
            return "N√©gatif" # Pour un ou deux pics n√©gatifs, sans positif
        elif p_plus > 0 and p_moins > 0 and p_plus > p_moins:
            return "Mixte positif"
        elif p_plus > 0 and p_moins > 0 and p_plus <= p_moins:
            return "Mixte tendu"
        # Si les deux sont √† 0, ou des √©tats invalides, classer comme "Aucune donn√©e"
        return "Aucune donn√©e"

# === INITIALISATION DES √âTATS STREAMLIT ======================================
default_states = {
    "etat": "menu",
    "participants": [],
    "services": [],
    "relations_saisies": [],
    "relation_a_modifier": None,
    "affiche_formulaire_participant": False,
    "participant_a_modifier": None,
    "selected_relations": pd.DataFrame(),
    "nombre_total_personnes": 0,
}
for k, v in default_states.items():
    st.session_state.setdefault(k, v)

# === UTILITAIRES D‚ÄôIMPORT / EXPORT ===========================================
def exporter_json_data() -> str:
    """Exporte les donn√©es actuelles de la session en une cha√Æne JSON."""
    data = {
        "participants": st.session_state.participants,
        "services": st.session_state.services,
        "relations_saisies": st.session_state.relations_saisies,
        "nombre_total_personnes": st.session_state.nombre_total_personnes,
    }
    return json.dumps(data, indent=4, ensure_ascii=False)


def exporter_excel_data() -> bytes:
    """
    Exporte les relations saisies dans un fichier Excel avec un ordre de colonnes sp√©cifique
    et ajuste automatiquement la largeur des colonnes. Ajoute √©galement une feuille de calcul
    avec un r√©capitulatif de la vigilance des relations, une feuille pour les relations
    unidirectionnelles et une feuille pour les relations crois√©es n√©gatives et positives.
    Inclut TOUTES les relations possibles (saisies et neutres, y compris celles
    impliquant des participants non nomm√©s) dans la feuille 'Relations'.
    """
    all_relations_for_excel = []

    # Construire la liste de TOUS les participants (nomm√©s + anonymes)
    all_person_names = [p['nom'] for p in st.session_state.participants]
    num_named_participants = len(all_person_names)
    total_persons_in_project = st.session_state.nombre_total_personnes

    if total_persons_in_project > num_named_participants:
        for i in range(total_persons_in_project - num_named_participants):
            all_person_names.append(f"Personne Anonyme {i+1}")

    # Cr√©er un dictionnaire de hachages pour les relations saisies pour une recherche rapide
    saisies_map = {}
    for rel in st.session_state.relations_saisies:
        key = (rel["√âmetteur"], rel["R√©cepteur"])
        saisies_map[key] = rel

    # D√©finir l'ordre exact des colonnes pour les feuilles Excel
    colonnes_relations_excel = [
        "√âmetteur", "R√©cepteur", "Date", "D√©but", "Fin", "Service",
        "P+", "P-", "I+", "I-", "C+", "C-",
        "Score Pic Positif", "Score Pic N√©gatif", "Score Net",
        "Vigilance", "Commentaire"
    ]

    # G√©n√©rer TOUTES les combinaisons bidirectionnelles possibles pour la feuille 'Relations'
    for emetteur in all_person_names:
        for recepteur in all_person_names:
            if emetteur != recepteur:
                current_key = (emetteur, recepteur)
                
                # V√©rifier si cette relation a √©t√© saisie
                if current_key in saisies_map:
                    all_relations_for_excel.append(saisies_map[current_key])
                else:
                    # C'est une relation neutre (non saisie)
                    # D√©terminer le service de l'√©metteur si l'√©metteur est nomm√©
                    service_emetteur = "RAS"
                    for p in st.session_state.participants:
                        if p["nom"] == emetteur:
                            service_emetteur = p["service"]
                            break

                    all_relations_for_excel.append({
                        "√âmetteur": emetteur,
                        "R√©cepteur": recepteur,
                        "Date": "RAS", # Pas de date pour une relation non saisie
                        "D√©but": "RAS", # Pas d'heure
                        "Fin": "RAS",   # Pas d'heure
                        "Service": service_emetteur, # Service de l'√©metteur (si nomm√©), sinon RAS
                        "P+": 0, "P-": 0, "I+": 0, "I-": 0, "C+": 0, "C-": 0,
                        "Score Pic Positif": 0,
                        "Score Pic N√©gatif": 0,
                        "Score Net": 0,
                        "Vigilance": "Neutre", # Marqu√© comme neutre
                        "Commentaire": "Relation non renseign√©e ou non applicable" # Commentaire explicatif
                    })
            
    df_relations = pd.DataFrame(all_relations_for_excel)

    # S'assurer que toutes les colonnes d√©finies existent dans le DataFrame
    for col in colonnes_relations_excel:
        if col not in df_relations.columns:
            df_relations[col] = None

    # R√©organiser le DataFrame selon l'ordre des colonnes d√©finies
    df_relations = df_relations[colonnes_relations_excel]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # --- Premi√®re feuille : Relations (Toutes les combinaisons, saisies et neutres) ---
        df_relations.to_excel(writer, index=False, sheet_name='Relations')

        workbook = writer.book
        worksheet_relations = writer.sheets['Relations']

        for col_idx, column_name in enumerate(colonnes_relations_excel):
            max_length = len(str(column_name))
            if not df_relations.empty:
                max_length = max(max_length, df_relations[column_name].astype(str).apply(len).max())

            adjusted_width = (max_length + 2)
            if column_name == "Commentaire":
                adjusted_width = min(adjusted_width, 100) # Limite pour le commentaire
            elif column_name in ["Date", "D√©but", "Fin", "Service", "Vigilance"]:
                 adjusted_width = min(adjusted_width, 25) # Limite pour ces colonnes
            else:
                 adjusted_width = min(adjusted_width, 15) # Limite g√©n√©rale pour les autres

            worksheet_relations.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = adjusted_width

        # --- Deuxi√®me feuille : R√©capitulatif Vigilance (R√©sum√© de la vigilance) ---
        worksheet_vigilance = workbook.create_sheet(title='R√©capitulatif Vigilance')

        # Calcul des statistiques de vigilance
        relations_enregistrees_count = len(st.session_state.relations_saisies)
        
        # Obtenir les comptages de chaque type de vigilance existant dans les relations SAISIES
        df_saisies_for_stats = pd.DataFrame(st.session_state.relations_saisies)
        vigilance_counts = df_saisies_for_stats['Vigilance'].value_counts().to_dict() if not df_saisies_for_stats.empty else {}
        
        # Initialiser les types de relations pour s'assurer qu'ils apparaissent m√™me s'ils sont √† 0
        types_relation_stats = ["Positif pur", "Positif", "Mixte positif", "Mixte tendu", "N√©gatif pur", "N√©gatif", "Aucune donn√©e"] # Updated types
        stats_data = []

        # Calcul du nombre total de combinaisons possibles (bas√© sur nombre_total_personnes)
        nombre_total_personnes_app = st.session_state.nombre_total_personnes 
        nombre_combinaisons_possibles = 0
        if nombre_total_personnes_app > 1:
            nombre_combinaisons_possibles = nombre_total_personnes_app * (nombre_total_personnes_app - 1)
        
        if nombre_combinaisons_possibles == 0 and nombre_total_personnes_app > 1:
            st.error("ERREUR DE CALCUL : Le nombre de combinaisons possibles est z√©ro. Veuillez saisir un nombre total de personnes sup√©rieur √† 1.")
        
        # Calcul des relations neutres GLOBALES (nombre total possible - nombre de relations enregistr√©es)
        relations_neutres_globales = max(0, nombre_combinaisons_possibles - relations_enregistrees_count)

        # Ajouter les relations sp√©cifiques (enregistr√©es)
        for rel_type in types_relation_stats:
            count = vigilance_counts.get(rel_type, 0)
            percentage = (count / nombre_combinaisons_possibles) if nombre_combinaisons_possibles > 0 else 0.0
            stats_data.append({"Type de relation": rel_type, "Nombre de cas": count, "Pourcentage": percentage})

        # Ajouter les relations neutres globales
        percentage_neutre_globale = (relations_neutres_globales / nombre_combinaisons_possibles) if nombre_combinaisons_possibles > 0 else 0.0
        stats_data.append({"Type de relation": "Neutre (Global)", "Nombre de cas": relations_neutres_globales, "Pourcentage": percentage_neutre_globale})
        
        # Ajouter la ligne "Total des combinaisons possibles" √† la fin
        total_pourcentage = 1.0 if nombre_combinaisons_possibles > 0 else 0.0
        stats_data.append({"Type de relation": "Total des combinaisons possibles", "Nombre de cas": nombre_combinaisons_possibles, "Pourcentage": total_pourcentage})


        # Cr√©er un DataFrame pour les statistiques
        df_stats = pd.DataFrame(stats_data)

        # √âcrire le DataFrame des stats dans la deuxi√®me feuille
        worksheet_vigilance.append(df_stats.columns.tolist())
        for row in df_stats.itertuples(index=False):
            worksheet_vigilance.append(list(row))

        # Ajuster la largeur des colonnes pour la feuille 'R√©capitulatif Vigilance'
        for col_idx, column_name in enumerate(df_stats.columns):
            max_length = len(str(column_name))
            if not df_stats.empty:
                max_length = max(max_length, df_stats[column_name].astype(str).apply(len).max())
            
            adjusted_width = (max_length + 2)
            if column_name == "Type de relation":
                adjusted_width = min(adjusted_width, 40)
            elif column_name == "Nombre de cas":
                adjusted_width = min(adjusted_width, 20)
            elif column_name == "Pourcentage":
                adjusted_width = min(adjusted_width, 20)
            
            worksheet_vigilance.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = adjusted_width
            
            # Format pour la colonne Pourcentage
            if column_name == "Pourcentage":
                for row_idx in range(2, worksheet_vigilance.max_row + 1):
                    cell = worksheet_vigilance.cell(row=row_idx, column=col_idx + 1)
                    cell.number_format = '0.00%'

        # --- Nouvelle feuille : Relations Unidirectionnelles ---
        worksheet_unidirectional = workbook.create_sheet(title='Relations Unidirectionnelles')
        
        # Filtrer relations_saisies pour inclure seulement les relations avec une vigilance d√©finie
        # (excluant 'Aucune donn√©e' qui r√©sulterait de P+=0 et P-=0)
        df_unidirectional = pd.DataFrame(st.session_state.relations_saisies)
        if not df_unidirectional.empty:
            df_unidirectional = df_unidirectional[df_unidirectional['Vigilance'] != 'Aucune donn√©e']
            # S'assurer de l'ordre des colonnes
            for col in colonnes_relations_excel:
                if col not in df_unidirectional.columns:
                    df_unidirectional[col] = None
            df_unidirectional = df_unidirectional[colonnes_relations_excel]
        
        df_unidirectional.to_excel(writer, index=False, sheet_name='Relations Unidirectionnelles')

        # Ajuster la largeur des colonnes pour la feuille 'Relations Unidirectionnelles'
        for col_idx, column_name in enumerate(colonnes_relations_excel):
            max_length = len(str(column_name))
            if not df_unidirectional.empty:
                max_length = max(max_length, df_unidirectional[column_name].astype(str).apply(len).max())

            adjusted_width = (max_length + 2)
            if column_name == "Commentaire":
                adjusted_width = min(adjusted_width, 100)
            elif column_name in ["Date", "D√©but", "Fin", "Service", "Vigilance"]:
                 adjusted_width = min(adjusted_width, 25)
            else:
                 adjusted_width = min(adjusted_width, 15)

            worksheet_unidirectional.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = adjusted_width

        # --- Nouvelle feuille : Relations Crois√©es N√©gatives ---
        worksheet_negative_cross = workbook.create_sheet(title='Relations Crois√©es N√©gatives')

        negative_cross_relations_data = []
        
        # Cr√©er un mapping pour une recherche rapide des relations saisies
        recorded_relations_lookup = {}
        for rel in st.session_state.relations_saisies:
            recorded_relations_lookup[(rel["√âmetteur"], rel["R√©cepteur"])] = rel

        # D√©finir les types de vigilance consid√©r√©s comme "n√©gatifs" pour cette analyse
        # NOTE: "Mixte tendu" indique un "pic n√©gatif" car p_moins >= p_plus
        # Ajout de "N√©gatif" aux types de vigilance n√©gatifs pour les relations crois√©es
        negative_vigilance_types_for_cross = {"N√©gatif pur", "N√©gatif", "Mixte tendu"}
        
        # It√©rer sur les paires uniques de participants nomm√©s pour trouver les relations crois√©es
        processed_pairs = set() # Pour √©viter de traiter (A,B) et (B,A) comme des paires diff√©rentes pour la logique
        for p1_obj in st.session_state.participants:
            for p2_obj in st.session_state.participants:
                p1 = p1_obj["nom"]
                p2 = p2_obj["nom"]
                
                if p1 == p2:
                    continue

                # S'assurer de traiter chaque paire une seule fois (ex: A,B mais pas B,A) pour la logique de paire
                pair_key = tuple(sorted((p1, p2)))
                if pair_key in processed_pairs:
                    continue
                processed_pairs.add(pair_key)

                # Obtenir les relations pour les deux directions
                rel_p1_to_p2 = recorded_relations_lookup.get((p1, p2))
                rel_p2_to_p1 = recorded_relations_lookup.get((p2, p1))

                # V√©rifier si les deux relations existent et sont d'un type de vigilance n√©gatif
                if (rel_p1_to_p2 and rel_p1_to_p2["Vigilance"] in negative_vigilance_types_for_cross and
                    rel_p2_to_p1 and rel_p2_to_p1["Vigilance"] in negative_vigilance_types_for_cross):
                    
                    # D√©terminer le type de relation crois√©e ("Conflit" ou "Tension relationnelle")
                    type_de_croise = ""
                    is_p1_p2_pure_negative = rel_p1_to_p2["Vigilance"] == "N√©gatif pur"
                    is_p2_p1_pure_negative = rel_p2_to_p1["Vigilance"] == "N√©gatif pur"

                    # "Conflit" si les deux sont "N√©gatif pur" (3 pics n√©gatifs de chaque c√¥t√©)
                    if is_p1_p2_pure_negative and is_p2_p1_pure_negative:
                        type_de_croise = "Conflit"
                    # "Tension relationnelle" si au moins l'un des deux est un "pic n√©gatif" (N√©gatif pur, N√©gatif ou Mixte tendu)
                    else:
                        type_de_croise = "Tension relationnelle"
                    
                    # Ajouter les deux relations √† la liste, avec la nouvelle classification
                    # Copier la relation pour √©viter de modifier l'objet original dans relations_saisies
                    rel_p1_to_p2_copy = rel_p1_to_p2.copy()
                    rel_p1_to_p2_copy["Type de Crois√©"] = type_de_croise
                    negative_cross_relations_data.append(rel_p1_to_p2_copy)

                    rel_p2_to_p1_copy = rel_p2_to_p1.copy()
                    rel_p2_to_p1_copy["Type de Crois√©"] = type_de_croise
                    negative_cross_relations_data.append(rel_p2_to_p1_copy)
        
        df_negative_cross = pd.DataFrame(negative_cross_relations_data)
        
        # Nouvel ordre de colonnes pour les relations crois√©es n√©gatives, incluant le type de crois√©
        colonnes_negative_cross_excel = colonnes_relations_excel + ["Type de Crois√©"]

        if not df_negative_cross.empty:
            # S'assurer que toutes les colonnes d√©finies existent dans le DataFrame
            for col in colonnes_negative_cross_excel:
                if col not in df_negative_cross.columns:
                    df_negative_cross[col] = None
            # R√©organiser le DataFrame selon l'ordre des colonnes d√©finies
            df_negative_cross = df_negative_cross[colonnes_negative_cross_excel]

        df_negative_cross.to_excel(writer, index=False, sheet_name='Relations Crois√©es N√©gatives')

        # Ajuster la largeur des colonnes pour la feuille 'Relations Crois√©es N√©gatives'
        for col_idx, column_name in enumerate(colonnes_negative_cross_excel):
            max_length = len(str(column_name))
            if not df_negative_cross.empty:
                max_length = max(max_length, df_negative_cross[column_name].astype(str).apply(len).max())

            adjusted_width = (max_length + 2)
            if column_name == "Commentaire":
                adjusted_width = min(adjusted_width, 100)
            elif column_name in ["Date", "D√©but", "Fin", "Service", "Vigilance", "Type de Crois√©"]:
                 adjusted_width = min(adjusted_width, 25)
            else:
                 adjusted_width = min(adjusted_width, 15)

            worksheet_negative_cross.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = adjusted_width


        # --- Nouvelle feuille : Relations Crois√©es Positives ---
        worksheet_positive_cross = workbook.create_sheet(title='Relations Crois√©es Positives')

        positive_cross_relations_data = []
        
        # D√©finir les types de vigilance consid√©r√©s comme "positifs" pour cette analyse
        positive_vigilance_types_for_cross = {"Positif pur", "Positif"}
        
        # It√©rer sur les paires uniques de participants nomm√©s pour trouver les relations crois√©es
        # Nous r√©utilisons processed_pairs pour √©viter les doublons A-B et B-A pour les relations positives
        processed_pairs_positive = set()
        for p1_obj in st.session_state.participants:
            for p2_obj in st.session_state.participants:
                p1 = p1_obj["nom"]
                p2 = p2_obj["nom"]
                
                if p1 == p2:
                    continue

                pair_key_positive = tuple(sorted((p1, p2)))
                if pair_key_positive in processed_pairs_positive:
                    continue
                processed_pairs_positive.add(pair_key_positive)

                # Obtenir les relations pour les deux directions
                rel_p1_to_p2 = recorded_relations_lookup.get((p1, p2))
                rel_p2_to_p1 = recorded_relations_lookup.get((p2, p1))

                # V√©rifier si les deux relations existent et sont d'un type de vigilance positif
                if (rel_p1_to_p2 and rel_p1_to_p2["Vigilance"] in positive_vigilance_types_for_cross and
                    rel_p2_to_p1 and rel_p2_to_p1["Vigilance"] in positive_vigilance_types_for_cross):
                    
                    # D√©terminer le type de relation crois√©e positive
                    type_de_croise = ""
                    is_p1_p2_pure_positive = rel_p1_to_p2["Vigilance"] == "Positif pur"
                    is_p2_p1_pure_positive = rel_p2_to_p1["Vigilance"] == "Positif pur"

                    # "Harmonie Parfaite" si les deux sont "Positif pur"
                    if is_p1_p2_pure_positive and is_p2_p1_pure_positive:
                        type_de_croise = "Harmonie Parfaite"
                    # "Harmonie Relationnelle" si au moins l'un des deux est "Positif" ou un mix avec "Positif pur"
                    else:
                        type_de_croise = "Harmonie Relationnelle"
                    
                    # Ajouter les deux relations √† la liste, avec la nouvelle classification
                    rel_p1_to_p2_copy = rel_p1_to_p2.copy()
                    rel_p1_to_p2_copy["Type de Crois√©"] = type_de_croise
                    positive_cross_relations_data.append(rel_p1_to_p2_copy)

                    rel_p2_to_p1_copy = rel_p2_to_p1.copy()
                    rel_p2_to_p1_copy["Type de Crois√©"] = type_de_croise
                    positive_cross_relations_data.append(rel_p2_to_p1_copy)
        
        df_positive_cross = pd.DataFrame(positive_cross_relations_data)
        
        # Nouvel ordre de colonnes pour les relations crois√©es positives, incluant le type de crois√©
        colonnes_positive_cross_excel = colonnes_relations_excel + ["Type de Crois√©"]

        if not df_positive_cross.empty:
            # S'assurer que toutes les colonnes d√©finies existent dans le DataFrame
            for col in colonnes_positive_cross_excel:
                if col not in df_positive_cross.columns:
                    df_positive_cross[col] = None
            # R√©organiser le DataFrame selon l'ordre des colonnes d√©finies
            df_positive_cross = df_positive_cross[colonnes_positive_cross_excel]

        df_positive_cross.to_excel(writer, index=False, sheet_name='Relations Crois√©es Positives')

        # Ajuster la largeur des colonnes pour la feuille 'Relations Crois√©es Positives'
        for col_idx, column_name in enumerate(colonnes_positive_cross_excel):
            max_length = len(str(column_name))
            if not df_positive_cross.empty:
                max_length = max(max_length, df_positive_cross[column_name].astype(str).apply(len).max())

            adjusted_width = (max_length + 2)
            if column_name == "Commentaire":
                adjusted_width = min(adjusted_width, 100)
            elif column_name in ["Date", "D√©but", "Fin", "Service", "Vigilance", "Type de Crois√©"]:
                 adjusted_width = min(adjusted_width, 25)
            else:
                 adjusted_width = min(adjusted_width, 15)

            worksheet_positive_cross.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = adjusted_width


        # --- Nouvelle feuille : R√©cap ---
        worksheet_recap = workbook.create_sheet(title='R√©cap')

        recap_dataframes = []

        if not df_unidirectional.empty:
            df_temp_uni = df_unidirectional.copy()
            df_temp_uni['Type de Crois√©'] = None  # Pas de type de crois√© pour les unidirectionnelles
            df_temp_uni['Type de R√©cap'] = 'Unidirectionnelle'
            recap_dataframes.append(df_temp_uni)
        
        if not df_negative_cross.empty:
            df_temp_neg = df_negative_cross.copy()
            df_temp_neg['Type de R√©cap'] = 'N√©gative Crois√©e'
            recap_dataframes.append(df_temp_neg)

        if not df_positive_cross.empty:
            df_temp_pos = df_positive_cross.copy()
            df_temp_pos['Type de R√©cap'] = 'Positive Crois√©e'
            recap_dataframes.append(df_temp_pos)
        
        df_recap = pd.DataFrame()
        if recap_dataframes:
            df_recap = pd.concat(recap_dataframes, ignore_index=True)
            
            # D√©finir l'ordre complet des colonnes pour le r√©capitulatif
            colonnes_recap_excel = colonnes_relations_excel + ["Type de Crois√©", "Type de R√©cap"]
            
            # S'assurer que toutes les colonnes d√©finies existent dans le DataFrame r√©capitulatif
            for col in colonnes_recap_excel:
                if col not in df_recap.columns:
                    df_recap[col] = None
            
            # R√©organiser le DataFrame selon l'ordre des colonnes d√©finies
            df_recap = df_recap[colonnes_recap_excel]


        df_recap.to_excel(writer, index=False, sheet_name='R√©cap')

        # Ajuster la largeur des colonnes pour la feuille 'R√©cap'
        for col_idx, column_name in enumerate(colonnes_recap_excel):
            max_length = len(str(column_name))
            if not df_recap.empty:
                max_length = max(max_length, df_recap[column_name].astype(str).apply(len).max())

            adjusted_width = (max_length + 2)
            if column_name == "Commentaire":
                adjusted_width = min(adjusted_width, 100)
            elif column_name in ["Date", "D√©but", "Fin", "Service", "Vigilance", "Type de Crois√©", "Type de R√©cap"]:
                 adjusted_width = min(adjusted_width, 25)
            else:
                 adjusted_width = min(adjusted_width, 15)

            worksheet_recap.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = adjusted_width


    output.seek(0)
    return output.getvalue()


def exporter_zip():
    """Cr√©e un fichier ZIP contenant le JSON et l'Excel des donn√©es du projet."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Ajouter le fichier JSON
        json_data = exporter_json_data()
        zf.writestr("barometre_projet.json", json_data)

        # Ajouter le fichier Excel
        excel_data = exporter_excel_data()
        zf.writestr("relations_barometre.xlsx", excel_data)

    zip_buffer.seek(0)
    return zip_buffer


def importer_json():
    st.markdown("### üìÅ Glissez-d√©posez un fichier JSON ici ou cliquez pour le s√©lectionner :")
    fichier = st.file_uploader("Charger un fichier JSON", type=["json"], label_visibility="collapsed")
    if fichier is not None:
        try:
            contenu = json.load(fichier)
            st.session_state.participants         = contenu.get("participants", [])
            st.session_state.services             = contenu.get("services", [])
            st.session_state.relations_saisies = contenu.get("relations_saisies", [])
            st.session_state.nombre_total_personnes = contenu.get("nombre_total_personnes", 0)
            st.success("Projet charg√© avec succ√®s !")
            st.session_state.etat = "relations"
            st.rerun()
        except Exception as e:
            st.error(f"Erreur lors du chargement du fichier JSON : {e}")


# === TITRE PRINCIPAL =========================================================
st.title("Barom√®tre Relationnel ‚Äî R√©alis√© par Hatice Gultekin")

# === MENU PRINCIPAL ==========================================================
if st.session_state.etat == "menu":
    st.subheader("Gestion de projet")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("D√©marrer un nouveau projet"):
            st.session_state.participants         = []
            st.session_state.services             = []
            st.session_state.relations_saisies = []
            st.session_state.relation_a_modifier = None
            st.session_state.selected_relations = pd.DataFrame()
            st.session_state.nombre_total_personnes = 0
            st.session_state.etat = "participants"
            st.rerun()
    # Appel direct de importer_json() en dehors de la colonne pour optimiser le glisser-d√©poser
    importer_json()

# === ETAPE 1 : PARTICIPANTS ==================================================
elif st.session_state.etat == "participants":
    st.subheader("√âtape 1 : Saisie des participants")

    # Formulaire d‚Äôajout
    with st.form("form_ajout_participant"):
        col1, col2 = st.columns(2)
        nom = col1.text_input("Nom du participant")
        nouveau_service = col2.text_input("Service associ√©")
        ajouter = st.form_submit_button("Ajouter")

    if ajouter:
        if nom.strip() and nouveau_service.strip():
            if all(nom != p["nom"] for p in st.session_state.participants):
                if nouveau_service not in st.session_state.services:
                    st.session_state.services.append(nouveau_service)
                st.session_state.participants.append(
                    {"nom": nom.strip(), "service": nouveau_service.strip()}
                )
                st.success(f"Participant ¬´ {nom.strip()} ¬ª ajout√©.")
                st.rerun()
            else:
                st.warning("Participant d√©j√† ajout√©.")
        else:
            st.warning("Veuillez saisir un nom ET un service valides.")

    if not st.session_state.participants:
        st.info("Ajoutez au moins deux participants pour continuer.")

    # S√©lecteur + bouton de modification
    if st.session_state.participants:
        noms_services = [f"{p['nom']} ({p['service']})"
                                 for p in st.session_state.participants]
        index_to_modify = st.selectbox(
            "Modifier un participant existant",
            options=noms_services,
            index=0 if noms_services else None
        )

        if index_to_modify and st.button("‚úèÔ∏è Modifier le participant", key="modifier_participant_menu"):
            st.session_state.participant_a_modifier = index_to_modify.split(" (")[0]
            st.rerun()

    if st.button("üóëÔ∏è Supprimer le participant", key="supprimer_participant_menu"):
        if index_to_modify:
            participant_nom = index_to_modify.split(" (")[0]
            st.session_state.participants = [
                p for p in st.session_state.participants if p["nom"] != participant_nom
            ]
            st.session_state.relations_saisies = [
                r for r in st.session_state.relations_saisies
                if r["√âmetteur"] != participant_nom and r["R√©cepteur"] != participant_nom
            ]
            st.success(f"Participant ¬´ {participant_nom} ¬ª et ses relations ont √©t√© supprim√©s.")
            st.rerun()
        else:
            st.warning("Veuillez s√©lectionner un participant √† supprimer.")


    # Formulaire de modification (√©tape 1)
    if st.session_state.participant_a_modifier:
        data = next((p for p in st.session_state.participants
                                 if p["nom"] == st.session_state.participant_a_modifier), None)
        if data:
            with st.form("modif_form_etape1"):
                new_nom = st.text_input("Nouveau nom", value=data["nom"])
                new_service = st.text_input("Nouveau service", value=data["service"])
                submit_modif = st.form_submit_button("Valider les modifications")

            if submit_modif:
                for rel in st.session_state.relations_saisies:
                    if rel["√âmetteur"] == data["nom"]:
                        rel["√âmetteur"] = new_nom.strip()
                    if rel["R√©cepteur"] == data["nom"]:
                        rel["R√©cepteur"] = new_nom.strip()
                data["nom"] = new_nom.strip()
                data["service"] = new_service.strip()
                st.success("Participant modifi√© avec succ√®s.")
                st.session_state.participant_a_modifier = None
                st.rerun()

    # Bouton suivant
    if len(st.session_state.participants) >= 2:
        if st.button("Passer √† l'√©tape suivante"):
            st.session_state.etat = "relations"
            st.rerun()
    elif st.session_state.participants:
        st.info("Ajoutez au moins un autre participant pour continuer.")

# === ETAPE 2 : RELATIONS =====================================================
elif st.session_state.etat == "relations":
    st.subheader("√âtape 2 : Saisie des relations")

    # --- Saisie du nombre total de personnes √† l'√©tape 2 ---
    st.markdown("---")
    st.markdown("### Nombre total de personnes concern√©es par l'analyse")
    with st.form("form_saisie_total_personnes"):
        current_total_personnes = st.session_state.nombre_total_personnes
        total_personnes_input = st.number_input(
            "Saisissez le nombre total de personnes (m√™me celles non d√©taill√©es ci-dessous) :",
            min_value=0,
            value=current_total_personnes,
            step=1,
            key="total_personnes_input_etape2"
        )
        ajouter_total_personnes = st.form_submit_button("Enregistrer le nombre de personnes")

    if ajouter_total_personnes:
        if total_personnes_input > 0:
            st.session_state.nombre_total_personnes = total_personnes_input
            st.success(f"Nombre total de personnes enregistr√© : {total_personnes_input}")
        else:
            st.warning("Veuillez saisir un nombre valide (> 0) de personnes.")
    st.markdown("---")


    # S√©lecteur + bouton de modification (√©tape 2)
    if st.session_state.participants:
        noms_services = [f"{p['nom']} ({p['service']})"
                                 for p in st.session_state.participants]
        index_to_modify_rel = st.selectbox(
            "Modifier un participant existant",
            options=noms_services,
            key="select_modifier_rel",
            index=0 if noms_services else None
        )
        if index_to_modify_rel and st.button("‚úèÔ∏è Modifier le participant", key="modifier_participant_relations"):
            st.session_state.participant_a_modifier = index_to_modify_rel.split(" (")[0]
            st.rerun()

        if st.button("üóëÔ∏è Supprimer le participant", key="supprimer_participant_menu"):
            if index_to_modify_rel:
                participant_nom = index_to_modify_rel.split(" (")[0]
                st.session_state.participants = [
                    p for p in st.session_state.participants if p["nom"] != participant_nom
                ]
                st.session_state.relations_saisies = [
                    r for r in st.session_state.relations_saisies
                    if r["√âmetteur"] != participant_nom and r["R√©cepteur"] != participant_nom
                ]
                st.success(f"Participant ¬´ {participant_nom} ¬ª et ses relations ont √©t√© supprim√©s.")
                st.rerun()
            else:
                st.warning("Veuillez s√©lectionner un participant √† supprimer.")


    # Formulaire de modification (affich√© dans l‚Äô√©tape 2)
    if st.session_state.participant_a_modifier:
        data = next((p for p in st.session_state.participants
                                 if p["nom"] == st.session_state.participant_a_modifier), None)
        if data:
            with st.form("modif_form_relations"):
                new_nom = st.text_input("Nouveau nom", value=data["nom"])
                new_service = st.text_input("Nouveau service", value=data["service"])
                submit_modif = st.form_submit_button("Valider les modifications")

            if submit_modif:
                for rel in st.session_state.relations_saisies:
                    if rel["√âmetteur"] == data["nom"]:
                        rel["√âmetteur"] = new_nom.strip()
                    if rel["R√©cepteur"] == data["nom"]:
                        rel["R√©cepteur"] = new_nom.strip()
                data["nom"] = new_nom.strip()
                data["service"] = new_service.strip()
                st.success("Participant modifi√© avec succ√®s.")
                st.session_state.participant_a_modifier = None
                st.rerun()

    # --- Ajout rapide d‚Äôun participant --------------------------------------
    with st.expander("‚ûï Ajouter un participant oubli√©"):
        with st.form("form_ajout_participant_rapide"):
            col1, col2 = st.columns(2)
            nom_rapide = col1.text_input("Nom du participant oubli√©")
            service_rapide = col2.text_input("Service associ√©")
            ajouter_rapide = st.form_submit_button("Ajouter")

        if ajouter_rapide:
            if nom_rapide.strip() and service_rapide.strip():
                if all(nom_rapide != p["nom"]
                                 for p in st.session_state.participants):
                    if service_rapide not in st.session_state.services:
                        st.session_state.services.append(service_rapide)
                    st.session_state.participants.append(
                        {"nom": nom_rapide.strip(),
                         "service": service_rapide.strip()}
                    )
                    st.success("Participant ajout√© avec succ√®s.")
                    st.rerun()
                else:
                    st.warning("Ce participant existe d√©j√†.")
            else:
                st.warning("Veuillez saisir un nom ET un service valides.")

    # --- Liste des relations possibles --------------------------------------
    noms = [p["nom"] for p in st.session_state.participants]
    relations_possibles = [(e, r) for e in noms for r in noms if e != r]
    relation_textes = [f"{i+1}. {e} ‚Üí {r}"
                                 for i, (e, r) in enumerate(relations_possibles)]

    if relations_possibles:
        current_selection_index = 0
        if "relation_choisie_index" in st.session_state and st.session_state.relation_choisie_index < len(relation_textes):
            current_selection_index = st.session_state.relation_choisie_index
        elif relation_textes:
            current_selection_index = 0
        else:
            current_selection_index = None

        relation_choisie = st.selectbox(
            "Relation √† enregistrer :",
            relation_textes,
            index=current_selection_index,
            key="relation_choisie_select"
        )

        if relation_choisie:
            st.session_state.relation_choisie_index = relation_textes.index(relation_choisie)


        if relation_choisie:
            try:
                _, rel_part = relation_choisie.split(". ", 1)
                emetteur, recepteur = rel_part.split(" ‚Üí ")
                service_emetteur = next(
                    (p["service"] for p in st.session_state.participants
                     if p["nom"] == emetteur), ""
                )

                # Formulaire d‚Äôenregistrement
                date_default = datetime.now().date()
                date  = st.date_input("Date", value=date_default)
                debut = st.text_input("Heure d√©but", key="heure_debut_input")
                fin   = st.text_input("Heure fin", key="heure_fin_input")
                st.markdown(f"**Service d√©tect√© automatiquement :** {service_emetteur}")

                indicateurs = ["P+", "P-", "I+", "I-", "C+", "C-"]
                indicateurs_values = {i: st.checkbox(i, key=f"indic_{i}_checkbox") for i in indicateurs}
                commentaire = st.text_area("Commentaire global", key="commentaire_input")

                if st.button("üíæ Enregistrer la relation"):
                    erreurs = []
                    if not debut.strip():
                        erreurs.append("Heure de d√©but manquante")
                    if not fin.strip():
                        erreurs.append("Heure de fin manquante")
                    if not any(indicateurs_values.values()):
                        erreurs.append("Aucun indicateur s√©lectionn√©")

                    is_duplicate = False
                    for rel in st.session_state.relations_saisies:
                        if (rel["√âmetteur"] == emetteur and
                            rel["R√©cepteur"] == recepteur and
                            rel["Date"] == date.strftime("%d/%m/%Y") and
                            rel["D√©but"] == debut and
                            rel["Fin"] == fin):
                            is_duplicate = True
                            break
                    if is_duplicate:
                        erreurs.append("Une relation avec le m√™me √©metteur, r√©cepteur, date, heure de d√©but et heure de fin existe d√©j√†.")


                    if erreurs:
                        for e in erreurs:
                            st.warning(e)
                    else:
                        p_plus  = sum(indicateurs_values[i]
                                      for i in ["P+", "I+", "C+"])
                        p_moins = sum(indicateurs_values[i]
                                      for i in ["P-", "I-", "C-"])
                        vigilance = AnalyseRelationnelle.classer_relation(
                            p_plus, p_moins
                        )

                        st.session_state.relations_saisies.append({
                            "√âmetteur": emetteur,
                            "R√©cepteur": recepteur,
                            "Date": date.strftime("%d/%m/%Y"),
                            "D√©but": debut,
                            "Fin": fin,
                            "Service": service_emetteur,
                            "P+": int(indicateurs_values["P+"]),
                            "P-": int(indicateurs_values["P-"]),
                            "I+": int(indicateurs_values["I+"]),
                            "I-": int(indicateurs_values["I-"]),
                            "C+": int(indicateurs_values["C+"]),
                            "C-": int(indicateurs_values["C-"]),
                            "Score Pic Positif": p_plus,
                            "Score Pic N√©gatif": p_moins,
                            "Score Net": p_plus - p_moins,
                            "Vigilance": vigilance,
                            "Commentaire": commentaire,
                        })
                        st.success("Relation enregistr√©e.")
                        st.rerun()
            except Exception as e:
                st.warning(f"Erreur lors de l'extraction des donn√©es de relation. Veuillez v√©rifier la s√©lection: {e}")
        else:
            st.warning("Aucune relation possible s√©lectionn√©e. Veuillez en ajouter.")
    else:
        st.warning("Aucune relation possible ; ajoutez plus de participants.")


    # --- Affichage des Relations dans AgGrid ---
    st.markdown("---")
    st.subheader("Relations enregistr√©es")
    if st.session_state.relations_saisies:
        colonnes_ordonnees = [
            "√âmetteur", "R√©cepteur", "Date", "D√©but", "Fin", "Service",
            "P+", "P-", "I+", "I-", "C+", "C-",
            "Score Pic Positif", "Score Pic N√©gatif", "Score Net",
            "Vigilance", "Commentaire"
        ]
        df = pd.DataFrame(st.session_state.relations_saisies)

        for col in colonnes_ordonnees:
            if col not in df.columns:
                df[col] = None

        df = df[colonnes_ordonnees]

        gb = GridOptionsBuilder.from_dataframe(df)
        gb.configure_selection(selection_mode="multiple", use_checkbox=True)

        # Configuration des colonnes pour AgGrid
        gb.configure_column("√âmetteur", header_name="√âmetteur", wrapText=True, autoHeight=True, minWidth=100)
        gb.configure_column("R√©cepteur", header_name="R√©cepteur", wrapText=True, autoHeight=True, minWidth=100)
        gb.configure_column("Date", header_name="Date", wrapText=True, autoHeight=True, minWidth=80)
        gb.configure_column("D√©but", header_name="D√©but", wrapText=True, autoHeight=True, minWidth=70)
        gb.configure_column("Fin", header_name="Fin", wrapText=True, autoHeight=True, minWidth=70)
        gb.configure_column("Service", header_name="Service", wrapText=True, autoHeight=True, minWidth=100)

        for col_name in ["P+", "P-", "I+", "I-", "C+", "C-", "Score Pic Positif", "Score Pic N√©gatif", "Score Net"]:
            gb.configure_column(col_name, header_name=col_name, type=["numericColumn", "numberColumnFilter", "customNumericFormat"], precision=0, minWidth=50)

        gb.configure_column("Vigilance", header_name="Vigilance", wrapText=True, autoHeight=True, minWidth=120)

        # Configuration sp√©cifique pour la colonne "Commentaire"
        gb.configure_column("Commentaire", header_name="Commentaire", wrapText=True, autoHeight=True, minWidth=200, flex=1)

        # Configuration de la grille globale
        gb.configure_grid_options(domLayout='normal')

        grid_options = gb.build()

        grid_response = AgGrid(
            df,
            gridOptions=grid_options,
            update_mode=GridUpdateMode.MODEL_CHANGED,
            data_return_mode=DataReturnMode.AS_INPUT,
            height=400,
            allow_unsafe_jscode=True,
            enable_enterprise_modules=False,
            key="grid_relations_display",
            columns_auto_size_mode=ColumnsAutoSizeMode.FIT_CONTENTS,
            js_code="""
            function(params) {
                params.api.resetRowHeights();
                params.api.autoSizeAllColumns();
            }
            """
        )

        if grid_response['selected_rows'] is not None and not grid_response['selected_rows'].empty:
            st.session_state.selected_relations = grid_response['selected_rows']
        else:
            st.session_state.selected_relations = pd.DataFrame()

        if st.button("üóëÔ∏è Supprimer les relations s√©lectionn√©es", key="delete_selected_relations_button"):
            if not st.session_state.selected_relations.empty:
                selected_ids = set()
                for s_row in st.session_state.selected_relations.to_dict(orient='records'):
                    # Cr√©e un hash bas√© sur les champs identifiants pour trouver les doublons uniques
                    row_hash = hashlib.md5(json.dumps({k: s_row[k] for k in ["√âmetteur", "R√©cepteur", "Date", "D√©but", "Fin"]}, sort_keys=True).encode('utf-8')).hexdigest()
                    selected_ids.add(row_hash)

                new_relations_saisies = []
                for r in st.session_state.relations_saisies:
                    current_row_hash = hashlib.md5(json.dumps({k: r[k] for k in ["√âmetteur", "R√©cepteur", "Date", "D√©but", "Fin"]}, sort_keys=True).encode('utf-8')).hexdigest()
                    if current_row_hash not in selected_ids:
                        new_relations_saisies.append(r)
                
                st.session_state.relations_saisies = new_relations_saisies
                st.success("Relations s√©lectionn√©es supprim√©es.")
                st.rerun()
            else:
                st.warning("Veuillez s√©lectionner des relations √† supprimer.")

    else:
        st.info("Aucune relation enregistr√©e pour l'instant. Saisissez de nouvelles relations ci-dessus.")

    # --- Boutons de navigation ---
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Retour au menu principal"):
            st.session_state.etat = "menu"
            st.rerun()
    with col2:
        # Bouton d'exportation pour l'ensemble du projet au format ZIP
        download_zip_filename = f"barometre_projet_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        st.download_button(
            label="üì¶ T√©l√©charger le projet (JSON + Excel)",
            data=exporter_zip(),
            file_name=download_zip_filename,
            mime="application/zip",
            key="download_project_zip_button"
        )
